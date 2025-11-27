from django.shortcuts import render, redirect
from django.db.models import Avg, Sum
from django.utils import timezone
from decimal import Decimal
from .models import Producto, LoteAves, ReporteDiarioAves, ReporteProduccion
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import timedelta
from django.contrib.auth.decorators import login_required
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@login_required
def dashboard(request):
    if request.user.groups.filter(name='Trabajadores').exists():
        return redirect('menu_trabajador')
    
    # 1. OBTENER FILTRO DE TIEMPO (Por defecto 30 días)
    dias_param = request.GET.get('dias', '30') # Si no hay parametro, usa 30
    dias = int(dias_param)
    
    fecha_inicio = timezone.now().date() - timedelta(days=dias)
    titulo_periodo = f"Últimos {dias} días"

    # --- 2. LÓGICA DE ALGAS ---
    productos = Producto.objects.all()
    nombres_prod = []
    stock_prod = []
    proyeccion_prod = []

    for p in productos:
        # ANTES: Estaba fijo en * 7
        # proyeccion = p.stock_actual + (p.capacidad_maxima_diaria * 7)
        
        # AHORA: Usamos la variable 'dias' que viene del botón (7, 30 o 365)
        proyeccion = p.stock_actual + (p.capacidad_maxima_diaria * dias)
        
        nombres_prod.append(p.nombre)
        stock_prod.append(float(p.stock_actual)) 
        proyeccion_prod.append(float(proyeccion))

    # --- 3. LÓGICA DE GALLINAS (Rendimiento en el PERIODO seleccionado) ---
    lotes_algas = LoteAves.objects.filter(tipo_dieta='ALGAS', activo=True)
    lotes_control = LoteAves.objects.filter(tipo_dieta='CONTROL', activo=True)

    def calcular_rendimiento(lotes, fecha_limite):
        total_huevos = 0
        total_aves_inicial = 0
        hay_datos = False

        for lote in lotes:
            # AQUI APLICAMOS EL FILTRO DE FECHA __gte (Mayor o igual a fecha_limite)
            reportes = ReporteDiarioAves.objects.filter(
                lote=lote, 
                fecha_reporte__gte=fecha_limite 
            )
            
            if reportes.exists():
                hay_datos = True
                # Sumamos huevos
                total_huevos += reportes.aggregate(Sum('huevos_recolectados'))['huevos_recolectados__sum']
                # Usamos una simplificación para el promedio:
                # (Total Huevos / Días Reportados) / Cantidad Aves
                dias_reportados = reportes.count()
                total_aves_inicial += (lote.cantidad_aves_inicial * dias_reportados)
        
        if total_aves_inicial > 0 and hay_datos:
            # El cálculo es: Total Huevos del periodo / (Aves * Días)
            # Esto nos da la "Tasa de Postura Diaria Promedio"
            tasa = (total_huevos / total_aves_inicial)
            return round(tasa * 100, 2) # Lo devolvemos en porcentaje (Ej: 85.5%)
        return 0

    rendimiento_algas = calcular_rendimiento(lotes_algas, fecha_inicio)
    rendimiento_control = calcular_rendimiento(lotes_control, fecha_inicio)

    ultimos_algas = ReporteProduccion.objects.select_related('producto').order_by('-fecha_registro', '-created_at')[:5]
    
    # Traemos los últimos 5 reportes de la Granja
    ultimos_granja = ReporteDiarioAves.objects.select_related('lote').order_by('-fecha_reporte', '-created_at')[:5]

    context = {
        'nombres_prod': nombres_prod,
        'stock_prod': stock_prod,
        'proyeccion_prod': proyeccion_prod,
        'rendimiento_algas': rendimiento_algas,
        'rendimiento_control': rendimiento_control,
        'ganador': 'ALGAS' if rendimiento_algas > rendimiento_control else 'CONTROL',
        
        # Enviamos datos extras para los botones del HTML
        'dias_actual': dias, 
        'titulo_periodo': titulo_periodo,

        'bitacora_algas': ultimos_algas,
        'bitacora_granja': ultimos_granja
    }

    return render(request, 'dashboard.html', context)

@login_required
def ingreso_algas(request):
    """
    Vista para el formulario de los trabajadores.
    Envía la lista de productos para que puedan elegir.
    """
    # Filtramos solo lo que sea "Materia Prima" o productos relevantes para producción
    productos = Producto.objects.all() 
    
    return render(request, 'ingreso_algas.html', {'productos': productos})


@csrf_exempt # Esto permite recibir datos sin el token de seguridad estricto (ideal para prototipos/APIs simples)
def api_guardar_produccion(request):
    if request.method == 'POST':
        try:
            # 1. Convertimos los datos que vienen de JavaScript a Python
            data = json.loads(request.body)
            
            # 2. Buscamos el producto por su ID
            producto_id = data.get('producto_id')
            cantidad = data.get('cantidad')
            responsable = data.get('responsable')
            fecha_offline = data.get('fecha_registro') # La fecha real cuando se hizo el trabajo
            
            prod_obj = Producto.objects.get(id=producto_id)
            
            # 3. Creamos el reporte en la Base de Datos
            nuevo_reporte = ReporteProduccion.objects.create(
                producto=prod_obj,
                cantidad_producida=cantidad,
                responsable=responsable,
                fecha_registro=fecha_offline if fecha_offline else timezone.now()
            )
            
            # 4. Actualizamos el stock automáticamente
            prod_obj.stock_actual += Decimal(str(cantidad))
            prod_obj.save()
            
            return JsonResponse({'status': 'ok', 'mensaje': 'Guardado exitosamente'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido'}, status=405)

@login_required
def ingreso_aves(request):
    """
    Muestra el formulario para el granjero.
    Solo mostramos los lotes que están 'activos'.
    """
    lotes = LoteAves.objects.filter(activo=True)
    return render(request, 'ingreso_aves.html', {'lotes': lotes})

@csrf_exempt
def api_guardar_aves(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Datos recibidos del formulario
            lote_id = data.get('lote_id')
            huevos = int(data.get('huevos')) # Entero
            alimento = data.get('alimento')  # Decimal/Float
            mortalidad = int(data.get('mortalidad', 0)) # Opcional, por defecto 0
            responsable = data.get('responsable')
            fecha_offline = data.get('fecha_registro')

            # Buscar el lote
            lote_obj = LoteAves.objects.get(id=lote_id)

            # Crear reporte
            ReporteDiarioAves.objects.create(
                lote=lote_obj,
                huevos_recolectados=huevos,
                alimento_consumido_kg=Decimal(str(alimento)), # <-- Usamos el truco del Decimal aquí también
                mortalidad=mortalidad,
                # Observaciones automáticas si es offline
                observaciones=f"Rep. por {responsable} (Offline)" if fecha_offline else f"Rep. por {responsable}",
                fecha_reporte=fecha_offline if fecha_offline else timezone.now()
            )

            return JsonResponse({'status': 'ok', 'mensaje': 'Datos de granja guardados'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)
            
    return JsonResponse({'status': 'error'}, status=405)

@login_required
def exportar_algas_csv(request):
    """Genera un Excel (.xlsx) estilizado y completo"""
    
    # 1. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Algas"

    # 2. Definir Estilos (Azul corporativo)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a252f", end_color="1a252f", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # 3. Escribir Cabecera
    headers = ['Fecha Registro', 'Fecha Sistema', 'Producto', 'Categoría', 'Cantidad (Kg/L)', 'Responsable']
    ws.append(headers)

    # Aplicar estilo a la cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 4. Escribir Datos
    reportes = ReporteProduccion.objects.all().select_related('producto').order_by('-fecha_registro')
    
    total_kg = 0
    for r in reportes:
        # Sumamos para el total final
        total_kg += float(r.cantidad_producida)
        
        ws.append([
            r.fecha_registro,
            r.created_at.strftime("%d/%m/%Y"), # Formato limpio
            r.producto.nombre,
            r.producto.get_categoria_display(),
            float(r.cantidad_producida), # Importante que sea número para que Excel sume
            r.responsable
        ])

    # 5. Agregar Fila de Totales al final
    ultima_fila = ws.max_row + 1
    ws.cell(row=ultima_fila, column=4, value="TOTAL PRODUCIDO:").font = Font(bold=True)
    ws.cell(row=ultima_fila, column=5, value=total_kg).font = Font(bold=True, color="008000")

    # 6. Ajustar ancho de columnas automático
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Ej: 'A', 'B'
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 7. Preparar respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Algas_Completo.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_granja_csv(request):
    """Genera un Excel (.xlsx) estilizado para Granja"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Granja"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="e65100", end_color="e65100", fill_type="solid") # Naranja para granja

    # Cabecera
    headers = ['Fecha Reporte', 'Lote', 'Dieta', 'Huevos', 'Alimento (Kg)', 'Mortalidad', 'Responsable', 'Obs']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Datos
    reportes = ReporteDiarioAves.objects.all().select_related('lote').order_by('-fecha_reporte')
    
    total_huevos = 0
    total_alimento = 0

    for r in reportes:
        total_huevos += r.huevos_recolectados
        total_alimento += float(r.alimento_consumido_kg)
        
        ws.append([
            r.fecha_reporte,
            r.lote.nombre,
            r.lote.get_tipo_dieta_display(),
            r.huevos_recolectados,
            float(r.alimento_consumido_kg),
            r.mortalidad,
            r.observaciones
        ])

    # Totales
    ultima_fila = ws.max_row + 1
    ws.cell(row=ultima_fila, column=3, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=ultima_fila, column=4, value=total_huevos).font = Font(bold=True)
    ws.cell(row=ultima_fila, column=5, value=total_alimento).font = Font(bold=True)

    # Auto-ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Granja_Completo.xlsx"'
    wb.save(response)
    return response


@login_required
def menu_trabajador(request):
    """
    Pantalla simple para que el trabajador elija qué reportar.
    """
    return render(request, 'menu_trabajador.html')